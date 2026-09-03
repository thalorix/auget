import re
import py_compile

with open("app.py", "r", encoding="utf-8") as f:
    content = f.read()

# 1. Aggiungi modello Scenario se non esiste
if 'class Scenario(db.Model):' not in content:
    match = re.search(r'(class Report\(db\.Model\):.*?)(\n\nclass |\n@app\.route)', content, re.DOTALL)
    if match:
        scenario_model = '''

class Scenario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    report_id = db.Column(db.Integer, db.ForeignKey('report.id'), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    revenue = db.Column(db.Float, nullable=False)
    ebit = db.Column(db.Float, nullable=False)
    total_debt = db.Column(db.Float, nullable=False)
    interest_rate = db.Column(db.Float, nullable=False)
    cash = db.Column(db.Float, nullable=False)
    scenario_type = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    report = db.relationship('Report', backref=db.backref('scenarios', lazy=True))
'''
        content = content[:match.end(1)] + scenario_model + content[match.start(2):]
        print("✅ Modello Scenario aggiunto")

# 2. Sostituisci la rotta /simula
simula_start = content.find('@app.route("/simula"')
if simula_start != -1:
    next_route_match = re.search(r'\n(@app\.route\(|def [a-zA-Z_]+\()', content[simula_start + 20:])
    simula_end = simula_start + 20 + next_route_match.start() if next_route_match else len(content)
    
    new_simula = '''
@app.route("/simula", methods=["GET", "POST"])
@login_required
def simula():
    try: db.create_all()
    except: pass
    rep = Report.query.filter_by(user_id=current_user.id).order_by(Report.created_at.desc()).first()
    if not rep:
        return render_template_string(BASE_TEMPLATE, title="Simula", content="<div class='card'><h2>Stress Test Pro</h2><p>Carica prima un bilancio in /analyze.</p><a href='/analyze' class='btn2'>Analizza</a></div>")
    
    m = _json.loads(rep.metrics_json or "{}")
    orig_rev = float(m.get("revenue") or 0)
    orig_ebit = float(m.get("ebit") or 0)
    orig_debt = float(m.get("total_debt") or 0)
    orig_cash = float(m.get("cassa") or 0)
    orig_int = float(m.get("interest") or 0)
    orig_ebitda = float(m.get("ebitda") or orig_ebit * 1.2)
    orig_rate = (orig_int / orig_debt * 100) if orig_debt > 0 else 5.0
    
    lev_map = {"Tech": 3.0, "Software": 3.5, "Services": 2.0, "Manufacturing": 2.0, "Retail": 1.3, "Finance": 1.0, "Altro": 1.8}
    sector = rep.sector or "Altro"
    lev = lev_map.get(sector, 1.8)
    
    scens = {
        "recession": {"name": "Recessione", "color": "#f97316", "rev": orig_rev*0.85, "ebit": orig_ebit*(1-0.15*lev), "debt": orig_debt, "rate": orig_rate+1.0, "cash": orig_cash},
        "crisis": {"name": "Crisi Grave", "color": "#ef4444", "rev": orig_rev*0.60, "ebit": orig_ebit*(1-0.40*lev), "debt": orig_debt*1.15, "rate": orig_rate+2.5, "cash": orig_cash*0.70},
        "growth": {"name": "Crescita", "color": "#10b981", "rev": orig_rev*1.25, "ebit": orig_ebit*(1+0.25*lev), "debt": orig_debt*1.30, "rate": orig_rate, "cash": orig_cash*0.80}
    }
    
    action = request.form.get("action", "calculate")
    if action == "save":
        nm = request.form.get("scenario_name", "Scenario").strip()
        if nm:
            db.session.add(Scenario(user_id=current_user.id, report_id=rep.id, name=nm, revenue=float(request.form.get("revenue", orig_rev)), ebit=float(request.form.get("ebit", orig_ebit)), total_debt=float(request.form.get("total_debt", orig_debt)), interest_rate=float(request.form.get("interest_rate", orig_rate)), cash=float(request.form.get("cassa", orig_cash)), scenario_type=request.form.get("scenario_type", "custom")))
            db.session.commit(); flash("Scenario salvato!", "success")
        return redirect("/simula")
    elif action == "delete":
        sc = Scenario.query.get(request.form.get("scenario_id"))
        if sc and sc.user_id == current_user.id: db.session.delete(sc); db.session.commit(); flash("Eliminato", "success")
        return redirect("/simula")
    elif action == "compare":
        ids = request.form.getlist("scenario_ids")
        if len(ids) >= 2:
            scs = Scenario.query.filter(Scenario.id.in_(ids), Scenario.user_id == current_user.id).all()
            html = "<div class='card' style='overflow-x:auto'><table style='width:100%;border-collapse:collapse'><tr style='background:var(--bg)'><th style='padding:1rem'>Metrica</th>"
            for c in scs: html += f"<th style='padding:1rem;color:var(--gold)'>{c.name}</th>"
            html += "</tr>"
            for lbl, fn in [("Ricavi", lambda c: f"{c.revenue:.1f}"), ("EBIT", lambda c: f"{c.ebit:.1f}"), ("Debito", lambda c: f"{c.total_debt:.1f}"), ("Debt/EBITDA", lambda c: f"{(c.total_debt/orig_ebitda if orig_ebitda>0 else 0):.1f}x")]:
                html += f"<tr><td style='padding:0.8rem'>{lbl}</td>"
                for c in scs: html += f"<td style='padding:0.8rem;text-align:center'>{fn(c)}</td>"
                html += "</tr>"
            html += "</table></div><div style='text-align:center;margin-top:1rem'><a href='/simula' class='btn2'>Torna</a></div>"
            return render_template_string(BASE_TEMPLATE, title="Confronto", content=html)
        flash("Seleziona almeno 2 scenari", "error"); return redirect("/simula")

    if request.method == "POST" and action == "calculate":
        s_rev = max(0, float(request.form.get("revenue", orig_rev)))
        s_ebit = max(0, float(request.form.get("ebit", orig_ebit)))
        s_debt = max(0, float(request.form.get("total_debt", orig_debt)))
        s_rate = max(0, min(50, float(request.form.get("interest_rate", orig_rate))))
        s_cash = max(0, float(request.form.get("cassa", orig_cash)))
        s_name = request.form.get("scenario_name", "Personalizzato")
    else:
        s_rev, s_ebit, s_debt, s_rate, s_cash, s_name = orig_rev, orig_ebit, orig_debt, orig_rate, orig_cash, "Originale"
    
    s_int = s_debt * (s_rate / 100)
    s_ebitda = s_ebit * 1.2
    bp = max(0, ((s_ebit - s_int) / s_ebit) * 100) if s_int > 0 and s_ebit > 0 else (100.0 if s_int == 0 else 0.0)
    ic = s_ebit / s_int if s_int > 0 else 999.0
    de = s_debt / s_ebitda if s_ebitda > 0 else 0
    dscr = s_ebit / s_int if s_int > 0 else 999
    cr = s_cash / max((s_rev - s_ebit) / 12, s_rev / 24) if max((s_rev - s_ebit) / 12, s_rev / 24) > 0 else 999
    
    if bp >= 40 and ic >= 3.0 and de <= 3.0: st, sc_col = "FORTEZZA", "#10b981"
    elif bp >= 20 and ic >= 1.5 and de <= 4.0: st, sc_col = "RESILIENTE", "#fbbf24"
    elif bp >= 5: st, sc_col = "FRAGILE", "#f97316"
    else: st, sc_col = "A RISCHIO", "#ef4444"
    
    bench_map = {"Tech": 45, "Software": 50, "Retail": 15, "Finance": 30, "Altro": 30}
    bench = bench_map.get(sector, 30)
    saved = Scenario.query.filter_by(user_id=current_user.id, report_id=rep.id).order_by(Scenario.created_at.desc()).all()
    
    html = "<div class='card' style='padding:2rem'><h2 style='color:var(--gold)'>Stress Test Pro</h2>"
    html += '<form method="post"><input type="hidden" name="action" value="calculate"><input type="hidden" name="scenario_name" id="sname" value="Personalizzato"><input type="hidden" name="scenario_type" id="stype" value="custom">'
    html += '<div style="display:grid;grid-template-columns:repeat(3, 1fr);gap:1rem;margin:1rem 0">'
    for k, v in scens.items():
        html += f'<button type="button" class="btn2" style="background:{v["color"]}" onclick="applyScen(\\'{k}\\', \\'{v["name"]}\\')">{v["name"]}</button>'
    html += '</div><div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(150px, 1fr));gap:1rem;margin:1rem 0">'
    for fid, lbl, val in [("revenue", "Ricavi", s_rev), ("ebit", "EBIT", s_ebit), ("total_debt", "Debito", s_debt), ("interest_rate", "Tasso %", s_rate), ("cassa", "Cassa", s_cash)]:
        html += f'<div><label>{lbl}</label><input type="number" id="{fid}" name="{fid}" value="{round(val,2)}" step="0.1" min="0" style="width:100%;padding:0.5rem"></div>'
    html += '</div><div style="display:flex;gap:1rem;margin-top:1rem">'
    html += '<button type="submit" class="btn2" style="background:var(--teal)">Ricalcola</button>'
    html += '<button type="button" class="btn2" style="background:var(--gold);color:#000" onclick="saveScen()">Salva</button>'
    html += f'<a href="/simula/export?rid={rep.id}" class="btn2" style="background:var(--blue)">Export Excel</a></div></form>'
    
    html += f'''<script>
    const scens = {{recession: {{revenue: {round(orig_rev*0.85,2)}, ebit: {round(orig_ebit*(1-0.15*lev),2)}, total_debt: {round(orig_debt,2)}, interest_rate: {round(orig_rate+1.0,2)}, cassa: {round(orig_cash,2)}}, crisis: {{revenue: {round(orig_rev*0.60,2)}, ebit: {round(orig_ebit*(1-0.40*lev),2)}, total_debt: {round(orig_debt*1.15,2)}, interest_rate: {round(orig_rate+2.5,2)}, cassa: {round(orig_cash*0.70,2)}}, growth: {{revenue: {round(orig_rev*1.25,2)}, ebit: {round(orig_ebit*(1+0.25*lev),2)}, total_debt: {round(orig_debt*1.30,2)}, interest_rate: {round(orig_rate,2)}, cassa: {round(orig_cash*0.80,2)}}}};
    function applyScen(k, n) {{
        for(let id in scens[k]) document.getElementById(id).value = scens[k][id];
        document.getElementById('sname').value = n; document.getElementById('stype').value = k;
    }}
    function saveScen() {{
        let n = prompt('Nome scenario:', document.getElementById('sname').value);
        if(n) {{ document.getElementById('sname').value = n; let f = document.querySelector('form'); let i = document.createElement('input'); i.type='hidden'; i.name='action'; i.value='save'; f.appendChild(i); f.submit(); }}
    }}
    </script>'''
    
    html += f'<div class="card" style="border:3px solid {sc_col};padding:2rem;margin-top:2rem;text-align:center"><h3 style="color:{sc_col};font-size:2rem">{st}</h3>'
    html += f'<div style="font-size:4rem;font-weight:900;color:{sc_col}">-{int(bp)}%</div><p>Crollo ricavi sopportabile</p>'
    html += '<div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(150px, 1fr));gap:1rem;margin-top:1rem">'
    for lbl, val, col in [("Interest Cov", f"{ic:.1f}x", "#10b981" if ic>=3 else "#ef4444"), ("Debt/EBITDA", f"{de:.1f}x", "#10b981" if de<=2 else "#ef4444"), ("DSCR", f"{dscr:.1f}x", "#10b981" if dscr>=1.5 else "#ef4444"), ("Cash Runway", f"{min(int(cr),99) if cr<999 else '99+'} mesi", "#10b981" if cr>=12 else "#ef4444")]:
        html += f'<div style="background:var(--bg);padding:1rem;border-radius:8px"><h4>{lbl}</h4><p style="font-size:1.5rem;font-weight:700;color:{col}">{val}</p></div>'
    html += '</div></div>'
    
    if saved:
        html += '<div class="card" style="margin-top:2rem;padding:1rem"><h4>Scenari Salvati</h4><form method="post"><input type="hidden" name="action" value="compare">'
        for sc in saved:
            html += f'<label style="display:flex;gap:0.5rem;margin:0.5rem 0"><input type="checkbox" name="scenario_ids" value="{sc.id}"> {sc.name} <button type="submit" name="action" value="delete" formaction="" onclick="this.form.action=\\'/simula\\';this.form.innerHTML+=\\'<input type=hidden name=scenario_id value={sc.id}>\\';this.form.submit();return false" style="background:red;color:white;border:none;padding:2px 6px;border-radius:4px;margin-left:auto">X</button></label>'
        html += '<button type="submit" class="btn2" style="width:100%;margin-top:1rem">Confronta Selezionati</button></form></div>'
    
    html += '<p style="text-align:center;margin-top:1rem;color:var(--muted)">Leva settore: ' + str(lev) + 'x</p>'
    return render_template_string(BASE_TEMPLATE, title="Stress Test Pro", content=html)

@app.route("/simula/export")
@login_required
def simula_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    import io
    rep = Report.query.filter_by(user_id=current_user.id).order_by(Report.created_at.desc()).first()
    if not rep: return redirect("/simula")
    m = _json.loads(rep.metrics_json or "{}")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stress Test"
    ws['A1'] = f"Stress Test - {rep.company or 'Report'}"; ws['A1'].font = Font(bold=True, size=14)
    for i, (lbl, val) in enumerate([("Ricavi", m.get("revenue",0)), ("EBIT", m.get("ebit",0)), ("Debito", m.get("total_debt",0)), ("Cassa", m.get("cassa",0))], start=3):
        ws[f'A{i}'], ws[f'B{i}'] = lbl, val
    scs = Scenario.query.filter_by(user_id=current_user.id, report_id=rep.id).all()
    if scs:
        ws2 = wb.create_sheet("Scenari")
        ws2['A1'], ws2['B1'], ws2['C1'] = "Nome", "Ricavi", "EBIT"
        for i, sc in enumerate(scs, start=2): ws2[f'A{i}'], ws2[f'B{i}'], ws2[f'C{i}'] = sc.name, sc.revenue, sc.ebit
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"stress_test_{rep.id}.xlsx")
'''
    content = content[:simula_start] + new_simula + content[simula_end:]
    print("✅ Rotte /simula e /simula/export sostituite")

with open("app.py", "w", encoding="utf-8") as f:
    f.write(content)

try:
    py_compile.compile("app.py", doraise=True)
    print("✅ Sintassi Python verificata con successo!")
except SyntaxError as e:
    print(f"❌ Errore di sintassi riga {e.lineno}: {e.msg}")
